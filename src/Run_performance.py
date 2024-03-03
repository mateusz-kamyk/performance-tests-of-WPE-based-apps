#!/usr/bin/env python3

import numpy as np
from prettytable import PrettyTable
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import misc
import Memory_cmd
import Reboot_performance
import iQIYI_startup
import iQIYI_video_preview_opening_time
import iQIYI_video_startup
import iQIYI_exit
import Zee5_startup
import Zee5_video_preview_opening_time
import Zee5_video_preview_trailer
import Zee5_video_startup
import Zee5_exit
import utils
from Memory_cmd import MemoryTest

misc.TestIntro()
misc.TestList()

(T1M_GFX, T1M_MAIN, T1M_RES, T1M_CPU, T1M_MemAvailable, T1M_Memfree, T1M_PSS, T1M_USS,
 T2M_GFX, T2M_MAIN, T2M_RES, T2M_CPU, T2M_MemAvailable, T2M_Memfree, T2M_PSS, T2M_USS,
 T3M_GFX, T3M_MAIN, T3M_RES, T3M_CPU, T3M_MemAvailable, T3M_Memfree, T3M_PSS, T3M_USS,
 T4M_GFX, T4M_MAIN, T4M_RES, T4M_CPU, T4M_MemAvailable, T4M_Memfree, T4M_PSS, T4M_USS,
 T5M_GFX, T5M_MAIN, T5M_RES, T5M_CPU, T5M_MemAvailable, T5M_Memfree, T5M_PSS, T5M_USS,
 T6M_GFX, T6M_MAIN, T6M_RES, T6M_CPU, T6M_MemAvailable, T6M_Memfree, T6M_PSS, T6M_USS,
 T7M_GFX, T7M_MAIN, T7M_RES, T7M_CPU, T7M_MemAvailable, T7M_Memfree, T7M_PSS, T7M_USS,
 T8M_GFX, T8M_MAIN, T8M_RES, T8M_CPU, T8M_MemAvailable, T8M_Memfree, T8M_PSS, T8M_USS,
 T9M_GFX, T9M_MAIN, T9M_RES, T9M_CPU, T9M_MemAvailable, T9M_Memfree, T9M_PSS, T9M_USS) = ([] for _ in range(72))


for no in range(1,4):
    utils.no = no
    print("##############Iteration:",no,"##############")
    misc.Hash()
    print("TEST 1T: Reboot duration")
    Reboot_performance.RebootPerformance()
    misc.Hash()
    print("TEST 1M: EPG memory usage after reboot")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T1M_GFX.append((no, Memory_cmd.result_GFX)) 
    T1M_MAIN.append((no, Memory_cmd.result_MAIN))
    T1M_RES.append((no, Memory_cmd.result_RES)) 
    T1M_CPU.append((no, Memory_cmd.result_CPU)) 
    T1M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T1M_Memfree.append((no, Memory_cmd.result_MemFree))
    T1M_PSS.append((no, Memory_cmd.result_PSS))
    T1M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 2T: iQIYI start-up duration")
    iQIYI_startup.iQIYI_startup()
    misc.Hash()
    print("TEST 2M: iQIYI memory usage in homepage")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T2M_GFX.append((no, Memory_cmd.result_GFX)) 
    T2M_MAIN.append((no, Memory_cmd.result_MAIN))
    T2M_RES.append((no, Memory_cmd.result_RES)) 
    T2M_CPU.append((no, Memory_cmd.result_CPU)) 
    T2M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T2M_Memfree.append((no, Memory_cmd.result_MemFree))
    T2M_PSS.append((no, Memory_cmd.result_PSS))
    T2M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 3T: iQIYI video preview opening time")
    iQIYI_video_preview_opening_time.iQIYI_video_preview()
    misc.Hash()
    print("TEST 3M: iQIYI memory usage in video preview")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T3M_GFX.append((no, Memory_cmd.result_GFX)) 
    T3M_MAIN.append((no, Memory_cmd.result_MAIN))
    T3M_RES.append((no, Memory_cmd.result_RES)) 
    T3M_CPU.append((no, Memory_cmd.result_CPU)) 
    T3M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T3M_Memfree.append((no, Memory_cmd.result_MemFree))
    T3M_PSS.append((no, Memory_cmd.result_PSS))
    T3M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 4T: iQIYI video start-up time")
    iQIYI_video_startup.iQIYI_video_startup()
    misc.Hash()
    print("TEST 4M: iQIYI memory usage during video playback")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T4M_GFX.append((no, Memory_cmd.result_GFX)) 
    T4M_MAIN.append((no, Memory_cmd.result_MAIN))
    T4M_RES.append((no, Memory_cmd.result_RES)) 
    T4M_CPU.append((no, Memory_cmd.result_CPU)) 
    T4M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T4M_Memfree.append((no, Memory_cmd.result_MemFree))
    T4M_PSS.append((no, Memory_cmd.result_PSS))
    T4M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 5T: iQIYI exit time")
    iQIYI_exit.iQIYI_exit_time()
    misc.Hash()
    print("TEST 5M: EPG memory usage after iQIYI exit")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T5M_GFX.append((no, Memory_cmd.result_GFX)) 
    T5M_MAIN.append((no, Memory_cmd.result_MAIN))
    T5M_RES.append((no, Memory_cmd.result_RES)) 
    T5M_CPU.append((no, Memory_cmd.result_CPU)) 
    T5M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T5M_Memfree.append((no, Memory_cmd.result_MemFree))
    T5M_PSS.append((no, Memory_cmd.result_PSS))
    T5M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 6T: Zee5 start-up duration")
    Zee5_startup.Zee5_startup()
    misc.Hash()
    print("TEST 6M: Zee5 memory usage in homepage")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T6M_GFX.append((no, Memory_cmd.result_GFX)) 
    T6M_MAIN.append((no, Memory_cmd.result_MAIN))
    T6M_RES.append((no, Memory_cmd.result_RES)) 
    T6M_CPU.append((no, Memory_cmd.result_CPU)) 
    T6M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T6M_Memfree.append((no, Memory_cmd.result_MemFree))
    T6M_PSS.append((no, Memory_cmd.result_PSS))
    T6M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 7T: Zee5 video preview opening time")
    Zee5_video_preview_opening_time.Zee5_video_preview()
    misc.Hash()
    print("TEST 8T: Zee5 video preview trailer launching time")
    Zee5_video_preview_trailer.Zee5_video_preview_trailer()
    misc.Hash()
    print("TEST 7M: Zee5 memory usage in video preview (during trailer playback)")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T7M_GFX.append((no, Memory_cmd.result_GFX)) 
    T7M_MAIN.append((no, Memory_cmd.result_MAIN))
    T7M_RES.append((no, Memory_cmd.result_RES)) 
    T7M_CPU.append((no, Memory_cmd.result_CPU)) 
    T7M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T7M_Memfree.append((no, Memory_cmd.result_MemFree))
    T7M_PSS.append((no, Memory_cmd.result_PSS))
    T7M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 9T: Zee5 video start-up time")
    Zee5_video_startup.Zee5_video_startup()
    misc.Hash()
    print("TEST 8M: Zee5 memory usage during video playback")
    result_GFX, result_MAIN, result_RES, result_CPU, result_MemAvailable, result_MemFree, result_PSS, result_USS = Memory_cmd.start_results()
    MemoryTest()
    T8M_GFX.append((no, Memory_cmd.result_GFX)) 
    T8M_MAIN.append((no, Memory_cmd.result_MAIN))
    T8M_RES.append((no, Memory_cmd.result_RES)) 
    T8M_CPU.append((no, Memory_cmd.result_CPU)) 
    T8M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T8M_Memfree.append((no, Memory_cmd.result_MemFree))
    T8M_PSS.append((no, Memory_cmd.result_PSS))
    T8M_USS.append((no, Memory_cmd.result_USS))
    misc.Hash()
    print("TEST 10T: Zee5 exit time")
    Zee5_exit.Zee5_exit_time()
    misc.Hash()
    print("TEST 9M: EPG memory usage after Zee5 exit")
    MemoryTest()
    T9M_GFX.append((no, Memory_cmd.result_GFX)) 
    T9M_MAIN.append((no, Memory_cmd.result_MAIN))
    T9M_RES.append((no, Memory_cmd.result_RES)) 
    T9M_CPU.append((no, Memory_cmd.result_CPU)) 
    T9M_MemAvailable.append((no, Memory_cmd.result_MemAvailable)) 
    T9M_Memfree.append((no, Memory_cmd.result_MemFree))
    T9M_PSS.append((no, Memory_cmd.result_PSS))
    T9M_USS.append((no, Memory_cmd.result_USS)) 
    misc.Hash()

def extract_numeric_parts_GFX(result_GFX):
    # Extract numeric parts from result and average calculation
    numeric_values_GFX_x = [float(value.split('%')[0]) for _, value in result_GFX]
    numeric_values_GFX_y = [float(value.split('%')[1]) for _, value in result_GFX]
    avg_GFX_x = sum(numeric_values_GFX_x) / len(result_GFX)
    avg_GFX_y = sum(numeric_values_GFX_y) / len(result_GFX)

    return numeric_values_GFX_x, numeric_values_GFX_y, avg_GFX_x, avg_GFX_y

def extract_numeric_parts_MAIN(result_MAIN):
    # Extract numeric parts from result and average calculation
    numeric_values_MAIN_x = [float(value.split('%')[0]) for _, value in result_MAIN]
    numeric_values_MAIN_y = [float(value.split('%')[1]) for _, value in result_MAIN]
    avg_MAIN_x = sum(numeric_values_MAIN_x) / len(result_MAIN)
    avg_MAIN_y = sum(numeric_values_MAIN_y) / len(result_MAIN)

    return numeric_values_MAIN_x, numeric_values_MAIN_y, avg_MAIN_x, avg_MAIN_y

def extract_numeric_parts_RES(result_RES):
    # Extract numeric parts from result and average calculation
    numeric_values_RES = [float(value) for _, value in result_RES]
    avg_RES = sum(numeric_values_RES) / len(result_RES)

    return numeric_values_RES, avg_RES

def extract_numeric_parts_CPU(result_CPU):
    # Extract numeric parts from result and average calculation
    numeric_values_CPU = [float(value) for _, value in result_CPU]
    avg_CPU = sum(numeric_values_CPU) / len(result_CPU)

    return numeric_values_CPU, avg_CPU

def extract_numeric_parts_MemAvailable(result_MemAvailable):
    # Extract numeric parts from result and average calculation
    numeric_values_MemAvailable = [float(value) for _, value in result_MemAvailable]
    avg_MemAvailable = sum(numeric_values_MemAvailable) / len(result_MemAvailable)

    return numeric_values_MemAvailable, avg_MemAvailable

def extract_numeric_parts_MemFree(result_MemFree):
    # Extract numeric parts from result and average calculation
    numeric_values_MemFree = [float(value) for _, value in result_MemFree]
    avg_MemFree = sum(numeric_values_MemFree) / len(result_MemFree)

    return numeric_values_MemFree, avg_MemFree

def extract_numeric_parts_PSS(result_PSS):
    # Extract numeric parts from result and average calculation
    numeric_values_PSS = [float(value) for _, value in result_PSS]
    avg_PSS = sum(numeric_values_PSS) / len(result_PSS)

    return numeric_values_PSS, avg_PSS

def extract_numeric_parts_USS(result_USS):
    # Extract numeric parts from result and average calculation
    numeric_values_USS = [float(value) for _, value in result_USS]
    avg_USS = sum(numeric_values_USS) / len(result_USS)

    return numeric_values_USS, avg_USS


def set_row_heights(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                max_cell_height = max([len(str(cell.value).split('\n')) for cell in row])
                sheet.row_dimensions[row[0].row].height = max_cell_height * 30 

def set_borders(sheet):
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            # Sprawdź, czy wiersz nie jest pusty
            if any(cell.value is not None for cell in row):
                for cell in row:
                    cell.border = thin_border

def set_alignment(sheet):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2):  
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def Summary_GFX():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_GFX),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_GFX),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_GFX),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_GFX),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_GFX),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_GFX),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_GFX),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_GFX),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_GFX),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_GFX_x, numeric_values_GFX_y,avg_GFX_x, avg_GFX_y = extract_numeric_parts_GFX(test_results)


            if( avg_GFX_x > 0 and avg_GFX_y > 0):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_GFX_x[result[0]-1]:.0f} {numeric_values_GFX_y[result[0]-1]:.0f}", f"{avg_GFX_x:.0f} {avg_GFX_y:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_GFX_x[result[0]-1]:.0f} {numeric_values_GFX_y[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1
    

    sheet.append([])
    GFX_header = ["GFX memory usage"]
    sheet.append(GFX_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result (used, peak)[%]", "Average (used, peak)[%]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("GFX memory usage:")
    print(table)


def Summary_MAIN():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_MAIN),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_MAIN),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_MAIN),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_MAIN),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_MAIN),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_MAIN),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_MAIN),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_MAIN),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_MAIN),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_MAIN_x, numeric_values_MAIN_y,avg_MAIN_x, avg_MAIN_y = extract_numeric_parts_GFX(test_results)


            if( avg_MAIN_x > 0 and avg_MAIN_y > 0):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_MAIN_x[result[0]-1]:.0f} {numeric_values_MAIN_y[result[0]-1]:.0f}", f"{avg_MAIN_x:.0f} {avg_MAIN_y:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_MAIN_x[result[0]-1]:.0f} {numeric_values_MAIN_y[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1
    

    sheet.append([])
    
    MAIN_header = ["MAIN memory usage"]
    sheet.append(MAIN_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result (used, peak)[%]", "Average (used, peak)[%]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("MAIN memory usage:")
    print(table)


def Summary_RES():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_RES),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_RES),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_RES),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_RES),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_RES),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_RES),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_RES),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_RES),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_RES),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_RES, avg_RES = extract_numeric_parts_RES(test_results)


            if( avg_RES > 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_RES[result[0]-1]:.0f}", f"{avg_RES:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_RES[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1
    

    sheet.append([])
    
    RES_header = ["RES memory usage"]
    sheet.append(RES_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [MB]", "Average [MB]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("RES memory usage:")
    print(table)


def Summary_CPU():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_CPU),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_CPU),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_CPU),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_CPU),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_CPU),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_CPU),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_CPU),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_CPU),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_CPU),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_CPU, avg_CPU = extract_numeric_parts_CPU(test_results)


            if( avg_CPU >= 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_CPU[result[0]-1]:.0f}", f"{avg_CPU:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_CPU[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1

    sheet.append([])
    
    CPU_header = ["CPU usage"]
    sheet.append(CPU_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [%]", "Average [%]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("CPU usage:")
    print(table)


def Summary_MemAvailable():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_MemAvailable),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_MemAvailable),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_MemAvailable),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_MemAvailable),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_MemAvailable),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_MemAvailable),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_MemAvailable),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_MemAvailable),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_MemAvailable),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_MemAvailable, avg_MemAvailable = extract_numeric_parts_MemAvailable(test_results)


            if( avg_MemAvailable > 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_MemAvailable[result[0]-1]:.0f}", f"{avg_MemAvailable:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_MemAvailable[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1

    sheet.append([])
    
    MemAvailable_header = ["MemAvailable"]
    sheet.append(MemAvailable_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [MB]", "Average [MB]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("MemAvailable:")
    print(table)

def Summary_MemFree():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_Memfree),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_Memfree),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_Memfree),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_Memfree),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_Memfree),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_Memfree),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_Memfree),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_Memfree),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_Memfree),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_MemFree, avg_MemFree = extract_numeric_parts_MemFree(test_results)


            if( avg_MemFree > 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_MemFree[result[0]-1]:.0f}", f"{avg_MemFree:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_MemFree[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1

    sheet.append([])
    
    MemFree_header = ["MemFree"]
    sheet.append(MemFree_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [MB]", "Average [MB]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("MemFree:")
    print(table)

def Summary_PSS():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_PSS),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_PSS),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_PSS),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_PSS),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_PSS),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_PSS),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_PSS),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_PSS),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_PSS),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_PSS, avg_PSS = extract_numeric_parts_PSS(test_results)


            if( avg_PSS > 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_PSS[result[0]-1]:.0f}", f"{avg_PSS:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_PSS[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1

    sheet.append([])
    
    PSS_header = ["Procrank PSS memory usage"]
    sheet.append(PSS_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [MB]", "Average [MB]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("Procrank Pss memory usage:")
    print(table)


def Summary_USS():
    test_cases = [
        ("TEST 1M", "EPG memory usage after reboot", T1M_USS),
        ("TEST 2M", "iQIYI memory usage in home page", T2M_USS),
        ("TEST 3M", "iQIYI memory usage in video preview", T3M_USS),
        ("TEST 4M", "iQIYI memory usage during video playback", T4M_USS),
        ("TEST 5M", "EPG memory usage after iQIYI exit", T5M_USS),
        ("TEST 6M", "Zee5 memory usage in homepage", T6M_USS),
        ("TEST 7M", "Zee5 memory usage in video preview (during trailer playback)", T7M_USS),
        ("TEST 8M", "Zee5 memory usage during video playback", T8M_USS),
        ("TEST 9M", "EPG memory usage after Zee5 exit", T9M_USS),
    ]

    
    table_data = []
    added_test_cases = set()
    test_averages = {}

    try:
        wb = openpyxl.load_workbook("test_results.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active


    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            numeric_values_USS, avg_USS = extract_numeric_parts_USS(test_results)


            if( avg_USS > 0 ):
                if key not in added_test_cases:
                    table_data.append([test_case, test_description, result[0], f"{numeric_values_USS[result[0]-1]:.0f}", f"{avg_USS:.0f}" ])
                    added_test_cases.add(key)
                else:
                    table_data.append(["", "", result[0],f"{numeric_values_USS[result[0]-1]:.0f}", "" ])
                if key not in test_averages:
                    test_averages[key] = [result[1], 1]
                else:
                    test_averages[key][0] += result[1]
                    test_averages[key][1] += 1

    sheet.append([])
    
    USS_header = ["Procrank Uss memory usage"]
    sheet.append(USS_header)
    
    header_row = ["Test case", "Test description", "Test number", "Result [MB]", "Average [MB]"]
    sheet.append(header_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(color="FFFFFF", size=14)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

    # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    excel_file = "test_results.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    print("Procrank Uss memory usage:")
    print(table)


def Summary_Time():

    def set_row_heights(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            max_cell_height = max([len(str(cell.value).split('\n')) for cell in row])
            sheet.row_dimensions[row[0].row].height = max_cell_height * 30  # Adjust the multiplier as needed

    def set_borders(sheet):
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border

    def set_alignment(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2):  # Adjust max_col to the number of columns to align
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_font_color(sheet):
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", size=14)

    test_cases = [
        ("TEST 1T", "Reboot duration", Reboot_performance.T1T),
        ("TEST 2T_A", "iQIYI application: start-up time (Black screen duration)", iQIYI_startup.T2T_A),
        ("TEST 2T_B", "iQIYI application: start-up time (Total start-up time)", iQIYI_startup.T2T_B),
        ("TEST 3T", "iQIYI application: video preview opening time", iQIYI_video_preview_opening_time.T3T),
        ("TEST 4T", "iQIYI application: video start-up time", iQIYI_video_startup.T4T),
        ("TEST 5T", "iQIYI application: exit time", iQIYI_exit.T5T),
        ("TEST 6T", "Zee5 application: start-up time", Zee5_startup.T6T),
        ("TEST 7T", "Zee5 application: video preview opening time", Zee5_video_preview_opening_time.T7T),
        ("TEST 8T", "Zee5 application: video preview trailer launching time", Zee5_video_preview_trailer.T8T),
        ("TEST 9T", "Zee5 application: video start-up time", Zee5_video_startup.T9T),
        ("TEST 10T", "Zee5 application: exit time", Zee5_exit.T10T),
    ]

    table_data = []
    added_test_cases = set()
    test_averages = {}

    for test_case, test_description, test_results in test_cases:
        key = (test_case, test_description)
        for result in test_results:
            if key not in added_test_cases:
                table_data.append([test_case, test_description, result[0], "{:.6f}".format(result[1]), ""])
                added_test_cases.add(key)
            else:
                table_data.append(["", "", result[0], "{:.6f}".format(result[1]), ""])
            if key not in test_averages:
                test_averages[key] = [result[1], 1]
            else:
                test_averages[key][0] += result[1]
                test_averages[key][1] += 1


    for key, value in test_averages.items():
        test_case, test_description = key
        average = value[0] / value[1]
        for row in table_data:
            if (row[0], row[1]) == key and row[2]:
                row[4] = "{:.6f}".format(average)
                break

    # Tworzenie tabeli PrettyTable
    table = PrettyTable()
    table.field_names = ["Test case", "Test description", "Test number", "Result", "Average"]

    for row in table_data:
        table.add_row(row)

    # Wyświetlanie tabeli
    print(table)

    wb = openpyxl.Workbook()
    sheet = wb.active

    header_row = ["Test case", "Test description", "Test number", "Result", "Average"]
    sheet.append(header_row)

    # Apply styling to headers
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="3465a4", end_color="3465a4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    for row in table_data:
        sheet.append(row)

        # Apply styling to data cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths based on content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    set_borders(sheet)

    set_row_heights(sheet)

    set_alignment(sheet)

    set_font_color(sheet)

    excel_file = "test_results_time.xlsx"
    wb.save(excel_file)
    print(f"Data saved into file: {excel_file}")

Summary_GFX()
Summary_MAIN()
Summary_RES()
Summary_CPU()
Summary_MemAvailable()
Summary_MemFree()
Summary_PSS()
Summary_USS

Summary_Time()


